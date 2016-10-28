import datetime
import multiprocessing as mp
import os
import random
import statistics as sta

import openpyxl as xlwt


class generation:
    def __init__(self):
        ## INITIALISATION VARIABLES
        self.fit = 100
        self.totstu = 1000  ## TOTAL NUMBER OF STUDENTS
        self.stuperroom = 5  ## STUDENTS PER ROOM
        self.stulist = []  ## MAIN ARRAY THAT STORES THE STUDENTS CHOICES
        self.li = []  ## USED FOR RANDOMLY GENERATING CHROMOSOMES
        self.chromelist = []  ## CHROMOSOME CONTAINER FOR THE GENERATION
        self.optrang = [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,
                        4]  ## RANGE OF CHOICES AVAILABLE TO EACH ATTRIBUTE LENGTH OF THIS LIST DECIDES THE NUMBER OF ATTRIBUTES PER STUDENT
        self.mainlist = []

    ## THE CHROMOSOME CLASS
    class chromo:
        def __init__(self, a):
            self.p = a  ## LIST THAT CONTAINS THE STUDENTS GROUPED TOGETHER
            self.fitness = 0

        def prnt(self):  ## PRINT FUNCTION
            print("persons: ", self.p, "   fitness : ", self.fitness)

        def calc_fitness(self, stulist):  ## FITNESS FUNCTION
            ## THE MEAN OF THE VARIANCE IN THE CHOICES OF EACH ATTRIBUTE
            ## OF EACH OF THE STUDENT IN THAT PARTICULAR GROUP
            data = []
            for i in range(len(stulist[0])):
                data1 = []
                for j in range(len(self.p)):
                    data1.append(stulist[self.p[j]][i])
                data.append(sta.variance(data1))
            self.fitness = sta.mean(data)

    ## FUNCTION FOR RANDOMLY GENERATING THE DATA
    def generate_data(self):
        for i in range(self.totstu):
            self.stulist.append([])
            self.li.append(i)
            for j in range(len(self.optrang)):
                self.stulist[i].append(random.randrange(self.optrang[j]))

    def calcu_fitness(self):
        sum = 0
        for val in self.chromelist:
            val.calc_fitness(self.stulist)
            sum = sum + val.fitness
        for val in self.mainlist:
            val.calc_fitness(self.stulist)
            sum = sum + val.fitness
        self.fit = sum / (len(self.chromelist) + len(self.mainlist))
        ##print(self.fit)

    ## FUNCTION FOR GENERATING THE INITIAL CHROMOSOMES
    def setinitstate(self):
        for i in range(int(self.totstu / self.stuperroom)):
            a = []
            for j in range(self.stuperroom):
                k = self.li[random.randrange(len(self.li))]
                self.li.remove(k)
                a.append(k)
            self.chromelist.append(self.chromo(a))
            self.chromelist[i].calc_fitness(self.stulist)  ## CALCULATE FITNESS WHILE CREATING THE CHROMOSOME

    ## FUNCTION FOR PRINTING THE CURRENT GENERATION
    def print_gen(self):
        for i in range(len(self.chromelist)):
            print("chromosome ", i, " :")
            self.chromelist[i].prnt()

    ## FUNCTION FOR PERFORMING THE CROSSOVER
    def crossover(self):
        chlist1 = list(self.chromelist)
        chlist = []
        chlist1.sort(key=lambda x: x.fitness, reverse=False)
        for ch in chlist1:
            if ch.fitness < 1:
                self.mainlist.append(ch)
                chlist1.remove(ch)
        prev = 0
        ##print(len(self.mainlist))
        for ch in chlist1:
            ch.fitness = ((ch.fitness - chlist1[0].fitness) / (chlist1[len(chlist1) - 1].fitness - chlist1[0].fitness))
        while len(chlist1) > 0:
            select = []
            if (len(chlist1) > 1):
                for i in range(2):
                    num = random.randrange(len(chlist1))
                    select.append(chlist1[num])
                    chlist1.pop(num)
                pocross = self.cross(select)
                for i in range(len(pocross)):
                    chlist.append(self.chromo(pocross[i]))
            else:
                chlist.append(chlist1[0])
                chlist1.pop()
        return chlist

    def cross(self, a):

        leng = len(a)
        p = []
        pocross = []
        for i in range(leng):
            pocross.append([])
            for val in a[i].p:
                p.append(val)

        while len(p) > 0:
            for i in range(leng):
                a = p[random.randrange(len(p))]
                p.remove(a)
                pocross[i].append(a)
        return pocross


## THIS IS THE DRIVER PROGRAM , THIS IS IMPLEMENTED AS FUNCTION TO HELP WITH THREADING
def driver(b):
    for u in range(1):
        g1 = generation()
        g1.generate_data()
        g1.setinitstate()
        wb = xlwt.Workbook()
        wb1 = xlwt.Workbook()
        wb1.create_sheet(title="test1")
        c_sh1 = wb1["test1"]
        wb.create_sheet(title="init")
        c_sh = wb["init"]
        for m in range(1, len(g1.chromelist) + 1):
            for n in range(1, g1.stuperroom + 1):
                c_sh.cell(row=len(g1.mainlist) + m, column=n, value=g1.chromelist[m - 1].p[n - 1])
            c_sh.cell(row=len(g1.mainlist) + m, column=g1.stuperroom + 1, value=g1.chromelist[m - 1].fitness)
        c_sh.cell(row=len(g1.chromelist) + 1, column=1, value=g1.fit)
        sheetno = 1
        for i in range(b):
            print(u, ",", i, ",pid:", os.getpid())
            g2 = generation()
            a = g1.crossover()
            g2.stulist = list(g1.stulist)
            g2.chromelist = a
            g2.mainlist = g1.mainlist
            g2.calcu_fitness()
            g1 = g2
            wb.create_sheet("Generation " + str(sheetno))
            c_sh = wb["Generation " + str(sheetno)]
            c_sh1.cell(row=sheetno, column=1, value=len(g2.mainlist))
            c_sh1.cell(row=sheetno, column=2, value=g2.fit)
            sheetno = sheetno + 1
            for m in range(1, len(g2.mainlist) + 1):
                for n in range(1, g2.stuperroom + 1):
                    c_sh.cell(row=m, column=n, value=g2.mainlist[m - 1].p[n - 1])
                c_sh.cell(row=m, column=g1.stuperroom + 1, value=g2.mainlist[m - 1].fitness)
            for m in range(1, len(g2.chromelist) + 1):
                for n in range(1, g2.stuperroom + 1):
                    c_sh.cell(row=len(g2.mainlist) + m, column=n, value=g2.chromelist[m - 1].p[n - 1])
                c_sh.cell(row=len(g2.mainlist) + m, column=g1.stuperroom + 1, value=g2.chromelist[m - 1].fitness)
            ##print(g2.fit)
            c_sh.cell(row=len(g2.mainlist) + len(g2.chromelist) + 1, column=1, value=g2.fit)
        wb.save("ga" + str(os.getpid()) + str(datetime.datetime.now().time().hour) + str(
            datetime.datetime.now().time().minute) + str(datetime.datetime.now().time().second) + ".xlsx")
        wb1.save("rate" + str(os.getpid()) + str(datetime.datetime.now().time().hour) + str(
            datetime.datetime.now().time().minute) + str(datetime.datetime.now().time().second) + ".xlsx")


##THREADING PROGRAM OPTIMISED FOR DUAL CORE
if __name__ == '__main__':
    with mp.Pool(processes=4) as pool:
        pool.map(driver, [2, 4, 5, 2, 3, 5, 6, 6, 6])
